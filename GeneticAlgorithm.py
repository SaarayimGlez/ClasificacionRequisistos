from Encoding import Encoding
from Decoding import *
from sklearn.model_selection import train_test_split
import pickle
import torch
from torch.utils.data import TensorDataset, DataLoader
from DistributedTraining import *
import numpy as np
from sklearn.preprocessing import LabelEncoder
import random
from copy import deepcopy
import os
import sys
import openpyxl
import json



class GeneticAlgorithm():
    
    
    def read_data(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        opcion = 4
        if opcion == 1:
            x_path = os.path.join(base_dir, 'x-no-2')
            y_path = os.path.join(base_dir, 'y-no-2')
        if opcion == 2:
            x_path = os.path.join(base_dir, 'x-no-0')
            y_path = os.path.join(base_dir, 'y-no-0')
        if opcion == 3:
            x_path = os.path.join(base_dir, 'x-no-1')
            y_path = os.path.join(base_dir, 'y-no-1')
        if opcion == 4:
            x_path = os.path.join(base_dir, 'x-no-4')
            y_path = os.path.join(base_dir, 'y-no-4')
        
        try:
            x = pickle.load(open(x_path, 'rb'))
            y = pickle.load(open(y_path, 'rb'))
        except FileNotFoundError as e:
            sys.exit(1)
        
        le = LabelEncoder()
        y = le.fit_transform(y)
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2, stratify=y, random_state=42)
        num_classes = len(np.unique(y))
    
        return x_train, x_test, y_train, y_test, num_classes
    
    def get_dataloader(self):
        batch_size=32
        x_train, x_test, y_train, y_test, clases = self.read_data()
        x_train = torch.tensor(x_train, dtype=torch.float32)
        x_test = torch.tensor(x_test, dtype=torch.float32)
        y_train = torch.tensor(y_train, dtype=torch.long)
        y_test = torch.tensor(y_test, dtype=torch.long)
    
        train_ds = TensorDataset(x_train, y_train)
        test_ds = TensorDataset(x_test, y_test)
        
        train_dl = DataLoader(train_ds, batch_size=batch_size, shuffle=True)
        test_dl = DataLoader(test_ds, batch_size=batch_size)

    
        return train_dl, test_dl, clases
    
    
    
    def create_population(self, number_population):
        pop = []
        encoding = Encoding()
        min_conv = 2
        max_conv = 5
        min_full = 1
        max_full = 4
        num_epochs = 10
        max_params = 2e6
        lr = 10 ** random.uniform(-5, -2)
        w = random.uniform(0.1, 1.0)
        loss_func = nn.NLLLoss(reduction = "sum")
        
        train_dl, test_dl, clases = self.get_dataloader()
        
        for i in range(number_population):
            e = encoding.encoding(min_conv,max_conv,min_full,max_full)
            network = decoding(e, clases)
            cnn1 = CNN(e, network[0], network[1], network[2])
            num, f, accuracy, params = training('1', 'cpu', cnn1, num_epochs, loss_func, train_dl, test_dl, lr, w, max_params)
            
            individuo = {
                'encoding': e,
                'fitness': f,
                'accuracy': accuracy,
                'params': params,
                'lr': lr,
                'w':w
                }
            
            pop.append(individuo)
            
        return pop
    
    def calcular_accuracy(self, poblacion):
        min_conv = 2
        max_conv = 5
        min_full = 1
        max_full = 4
        num_epochs = 10
        max_params = 2e6
        lr = 1e-4
        w = 0.3
        loss_func = nn.NLLLoss(reduction = "sum")
        train_dl, test_dl, clases = self.get_dataloader()
        for individuo in poblacion:
            network = decoding(individuo['encoding'], clases)
            cnn1 = CNN(individuo['encoding'], network[0], network[1], network[2])
            num, f, accuracy, params = training('1', 'cpu', cnn1, num_epochs, loss_func, train_dl, test_dl, individuo['lr'], individuo['w'], max_params)
                
            individuo['fitness'] = f
            individuo['accuracy'] = accuracy

            
        return poblacion
    
    def calcular_valor_esperado(self, poblacion, numero_poblacion):
        nueva_poblacion = []
        total_aptitudes = 0
        suma = 0
        
        for individuo in poblacion:
            total_aptitudes = total_aptitudes + individuo['accuracy']
        
        total_aptitudes = total_aptitudes/numero_poblacion
        
        for individuo in poblacion:
            aptitud_i = individuo['accuracy'] / total_aptitudes
            nueva_poblacion.append((individuo, aptitud_i))
            suma = suma + aptitud_i
            
        return nueva_poblacion

    
    
    def seleccionar_padres_universal_estocastica(self, poblacion, numero_padres):
        padres = []
        suma = 0
        ptr = random.uniform(0, 1)
        
        poblacion_ve = self.calcular_valor_esperado(poblacion, numero_padres)
        
        for individuo, valor_esperado in poblacion_ve:
            suma = suma + valor_esperado
            while suma >= ptr:
                if suma>= ptr:
                    padres.append(individuo)
                ptr = ptr +1
        
        return padres
        
    
    def mutation(self, individuo, prob_mut):
        n_conv, n_full, first_level, second_level = individuo['encoding']
    

        if random.random() < prob_mut:
            delta = random.choice([-1, 1])
            n_conv = max(2, n_conv + delta)
            n_conv = min(n_conv, 6)
    
        total_layers = n_conv + n_full
        n_full = total_layers - n_conv
    
        while len(first_level) < total_layers:
            if len(first_level) < n_conv:
                # Añadir capa conv válida
                first_level.append({
                    'type': 'conv',
                    'nfilters': random.choice([4, 8, 16]),
                    'fsize': random.choice([3, 5, 7]),
                    'pool': random.choice(['max', 'avg', 'off']),
                    'psize': random.choice([2, 3])
                })
            else:
                first_level.append({
                    'type': 'fc',
                    'neurons': random.choice([4, 8, 16, 32, 64, 128])
                })

        first_level = first_level[:total_layers]
    
        for i in range(n_conv):
            capa = first_level[i]
            if capa.get('type') != 'conv':
                first_level[i] = {
                    'type': 'conv',
                    'nfilters': random.choice([4, 8, 16]),
                    'fsize': random.choice([3, 5, 7]),
                    'pool': random.choice(['max', 'avg', 'off']),
                    'psize': random.choice([2, 3])
                }
    
        for i in range(n_conv, n_conv + n_full):
            capa = first_level[i]
            if capa.get('type') != 'fc' or 'neurons' not in capa:
                first_level[i] = {
                    'type': 'fc',
                    'neurons': random.choice([4, 8, 16, 32, 64, 128])
                }
    
        # Mutación en capas convolucionales
        for i in range(n_conv):
            capa = first_level[i]
            if random.random() < prob_mut:
                capa['nfilters'] = random.choice([4, 8, 16])
            if random.random() < prob_mut:
                capa['fsize'] = random.choice([3, 5, 7])
            if random.random() < prob_mut:
                capa['pool'] = random.choice(['max', 'avg', 'off'])
            if random.random() < prob_mut:
                capa['psize'] = random.choice([2, 3])
    
        # Mutación en capas fully connected
        for i in range(n_conv, n_conv + n_full):
            capa = first_level[i]
            if random.random() < prob_mut:
                capa['neurons'] = random.choice([4, 8, 16, 32, 64, 128])
    
        # Mutación en conexiones residuales
        for i in range(len(second_level)):
            if random.random() < prob_mut:
                second_level[i] = 1 - second_level[i] 
    

        individuo['encoding'] = (n_conv, n_full, first_level, second_level)
    
        return individuo





    

    def sbx_crossover(self, padre1, padre2):
        eta=10
        bounds=None
        
        hijo1 = {}
        hijo2 = {}
    
        for key in ['lr', 'w']:
            x1 = padre1[key]
            x2 = padre2[key]
            if x1 > x2:
                x1, x2 = x2, x1
            
            if abs(x1 - x2) < 1e-14:
                hijo1[key] = x1
                hijo2[key] = x2
                continue
    

            xl, xu = bounds[key]
    
            rand = random.random()
            beta = 1.0 + (2.0 * (x1 - xl) / (x2 - x1))
            alpha = 2.0 - pow(beta, -(eta + 1))
            if rand <= 1.0 / alpha:
                beta_q = pow(rand * alpha, 1.0 / (eta + 1))
            else:
                beta_q = pow(1.0 / (2.0 - rand * alpha), 1.0 / (eta + 1))
            child1 = 0.5 * ((x1 + x2) - beta_q * (x2 - x1))
    
            beta = 1.0 + (2.0 * (xu - x2) / (x2 - x1))
            alpha = 2.0 - pow(beta, -(eta + 1))
            if rand <= 1.0 / alpha:
                beta_q = pow(rand * alpha, 1.0 / (eta + 1))
            else:
                beta_q = pow(1.0 / (2.0 - rand * alpha), 1.0 / (eta + 1))
            child2 = 0.5 * ((x1 + x2) + beta_q * (x2 - x1))
    

            hijo1[key] = min(max(child1, xl), xu)
            hijo2[key] = min(max(child2, xl), xu)
    
        return hijo1, hijo2



    def modular_crossover(self, padre1, padre2):

        h1 = deepcopy(padre1)
        h2 = deepcopy(padre2)

        n_conv1, n_full1, first_level1, second_level1 = h1['encoding']
        n_conv2, n_full2, first_level2, second_level2 = h2['encoding']
    
        # --- Cruza convolucional ---
        min_conv = min(n_conv1, n_conv2)
        for i in range(min_conv):
            if random.random() < 0.5:
                first_level1[i], first_level2[i] = first_level2[i], first_level1[i]
    
        # --- Cruza fully connected ---
        min_full = min(n_full1, n_full2)
        start1 = n_conv1
        start2 = n_conv2
        for i in range(min_full):
            if random.random() < 0.5:
                first_level1[start1 + i], first_level2[start2 + i] = first_level2[start2 + i], first_level1[start1 + i]
    
        # --- Cruza conexiones residuales ---
        min_len = min(len(second_level1), len(second_level2))
        for i in range(min_len):
            if random.random() < 0.5:
                second_level1[i], second_level2[i] = second_level2[i], second_level1[i]

        h1['encoding'] = (n_conv1, n_full1, first_level1, second_level1)
        h2['encoding'] = (n_conv2, n_full2, first_level2, second_level2)
    
        return h1, h2



    def sanitizar_encoding(self, encoding):
        n_conv, n_full, first_level, second_level = encoding
        for i, capa in enumerate(first_level):
            if capa['type'] == 'conv':
                if 'nfilters' not in capa:
                    capa['nfilters'] = random.choice([4, 8, 16])
                if 'fsize' not in capa:
                    capa['fsize'] = random.choice([3, 5, 7])
                if 'pool' not in capa:
                    capa['pool'] = random.choice(['off', 'avg', 'max'])
                if capa['pool'] in ['avg', 'max'] and 'psize' not in capa:
                    capa['psize'] = random.choice([2, 3])
            elif capa['type'] == 'fc':
                if 'neurons' not in capa:
                    capa['neurons'] = random.choice([4, 8, 16, 32, 64, 128])
        return (n_conv, n_full, first_level, second_level)

            
        
    def cruza_mutacion(self, padres, probabilidad_cruza, probabilidad_mutacion):
        nueva_poblacion = []
        resultado = []
    
        for i in range(0, len(padres), 2):
            #print(i)
            
            if i == len(padres) - 1:
                padre1 = padres[i]
                padre2 = padres[i - 1]
            else:
                padre1 = padres[i]
                padre2 = padres[i + 1]
    
            if random.random() < probabilidad_cruza:
                hijo1, hijo2 = self.modular_crossover(padre1, padre2)
                hiper1, hiper2 = self.sbx_crossover(padre1, padre2)
    
                hijo1['lr'] = hiper1['lr']
                hijo1['w'] = hiper1['w']
                hijo2['lr'] = hiper2['lr']
                hijo2['w'] = hiper2['w']
                
                
                hijo1['encoding'] = self.sanitizar_encoding(hijo1['encoding'])
                hijo2['encoding'] = self.sanitizar_encoding(hijo2['encoding'])
                
            else:
                hijo1 = deepcopy(padre1)
                hijo2 = deepcopy(padre2)
    
            nueva_poblacion.append(hijo1)
            nueva_poblacion.append(hijo2)
    
        for individuo in nueva_poblacion:
            if random.random() < probabilidad_mutacion:
                individuo_mutado = self.mutation(individuo, probabilidad_mutacion)
                individuo_mutado['encoding'] = self.sanitizar_encoding(individuo_mutado['encoding'])
                resultado.append(individuo_mutado)
            else:
                resultado.append(individuo)
    
        return resultado


    def reemplazar_peor_nuevo_con_mejor_original(self, poblacion_original, nueva_poblacion):

        mejor_original = max(poblacion_original, key=lambda ind: ind['accuracy'])
        

        idx_peor_nuevo = min(range(len(nueva_poblacion)), key=lambda i: nueva_poblacion[i]['accuracy'])

        nueva_poblacion[idx_peor_nuevo] = mejor_original
        
        return nueva_poblacion

            
    
    def algoritmo_genetico(self, numero_poblacion, p_cruza, p_mutam, n_generaciones):
        poblacion = self.create_population(numero_poblacion)

        generacion = 1
        
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Convergencia"
        ws["A1"] = "Aptitud"
        ws["B1"] = "Parámetros"
        
        mejor_individuo = max(poblacion, key=lambda ind: ind['accuracy'])
        mejor_accuracy_final = mejor_individuo['accuracy']
        mejor_encoding = mejor_individuo['encoding']
        mejor_fitness = mejor_individuo.get('fitness', None)
        mejor_lr = mejor_individuo.get('lr', None)
        mejor_w = mejor_individuo.get('w', None)
        ws[f"A{generacion+1}"] = mejor_accuracy_final
        resultados = {
            "accuracy": mejor_accuracy_final,
            "fitness": mejor_fitness,
            "lr": mejor_lr,
            "w": mejor_w,
            "encoding": mejor_encoding
        }
        
        ws[f"B{generacion+1}"] = json.dumps(resultados)
        
        print("Mejor accuracy: ", mejor_accuracy_final)
        print("---------------------------------")
        
        while generacion <= n_generaciones:
            generacion = generacion + 1 
            padres = self.seleccionar_padres_universal_estocastica(poblacion, numero_poblacion)

            
            nueva_poblacion = self.cruza_mutacion(padres, p_cruza, p_mutam)
            poblacion_f = self.calcular_accuracy(nueva_poblacion)
            
            poblacion = self.reemplazar_peor_nuevo_con_mejor_original(poblacion, poblacion_f)
            
            print ('GENERACION ', generacion)
            
            mejor_individuo = max(poblacion, key=lambda ind: ind['accuracy'])
            mejor_accuracy_final = mejor_individuo['accuracy']
            mejor_encoding = mejor_individuo['encoding']
            mejor_fitness = mejor_individuo.get('fitness', None)
            mejor_lr = mejor_individuo.get('lr', None)
            mejor_w = mejor_individuo.get('w', None)
            ws[f"A{generacion+1}"] = mejor_accuracy_final
            resultados = {
                "accuracy": mejor_accuracy_final,
                "fitness": mejor_fitness,
                "lr": mejor_lr,
                "w": mejor_w,
                "encoding": mejor_encoding
            }
            
            ws[f"B{generacion+1}"] = json.dumps(resultados)
            
            print("Mejor accuracy: ", mejor_accuracy_final)
            print("---------------------------------")
            
            
            
            
        wb.save("convergencia_no_4_3.xlsx")

                
ga = GeneticAlgorithm()
ga.algoritmo_genetico(38, 0.7722, 0.1578, 43)

